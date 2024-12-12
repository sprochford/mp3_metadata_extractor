import os  # Import the os module for interacting with the operating system (e.g., listing files)
import pandas as pd  # Import the pandas library for data manipulation and analysis (we'll use it for creating DataFrames)
from tinytag import TinyTag  # Import the TinyTag library for reading metadata from audio files
from openpyxl import load_workbook  # Import the load_workbook function from openpyxl for working with Excel files
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side  # Import styling classes from openpyxl

def extract_mp3_metadata(folder_path, csv_path, excel_path):
    """
    Extracts metadata from MP3 files, groups by album, 
    and saves to an Excel spreadsheet with:
      - Separate sheets for each album, sorted by track number.
      - A "All Songs" sheet with all tracks, sorted by album and track.
    Also saves all data to a CSV file.
    Applies formatting to the Excel spreadsheet, including grids.

    Args:
        folder_path (str): Path to the folder containing MP3s.
        csv_path (str): Path to save the CSV file.
        excel_path (str): Path to save the Excel file.
    """

    data = []  # Initialize an empty list to store the extracted metadata for each MP3 file

    # Loop through each item (file or folder) in the specified folder
    for filename in os.listdir(folder_path):
        if filename.endswith('.mp3'):  # Check if the item is an MP3 file (ends with '.mp3')
            file_path = os.path.join(folder_path, filename)  # Create the full path to the MP3 file

            tag = TinyTag.get(file_path)  # Use TinyTag to read the metadata (tags) from the MP3 file

            # Format duration in MM:SS
            duration_seconds = int(tag.duration)  # Get duration in seconds as integer
            minutes = duration_seconds // 60  # Calculate minutes
            seconds = duration_seconds % 60  # Calculate remaining seconds
            duration_formatted = f"{minutes:02d}:{seconds:02d}"  # Format as MM:SS

            # Create a dictionary to store the metadata for the current track
            track_info = {
                'Artist': tag.artist,  # Extract and store the artist tag
                'Title': tag.title,  # Extract and store the title tag
                'Duration': duration_formatted,  # Store the formatted duration in MM:SS
                'Album': tag.album,  # Extract and store the album tag
                'Track': tag.track,  # Extract and store the track number
                'Genre': tag.genre,  # Extract and store the genre tag
                'Filename': filename,  # Store the filename
                'Comment': tag.comment  # Extract and store the comment tag
            }
            data.append(track_info)  # Add the track_info dictionary to the data list

    # Create a pandas DataFrame from the collected metadata (data list)
    df = pd.DataFrame(data)

    # Save all data to CSV
    df.to_csv(csv_path, index=False)  # Save the DataFrame to a CSV file (index=False prevents row numbers from being saved)
    print(f"MP3 metadata extracted and saved to '{csv_path}'")  # Print a message indicating the CSV file has been created

    # Group data by album for separate sheets
    album_groups = df.groupby('Album')  # Group the DataFrame by the 'Album' column

    # Create an ExcelWriter object to write to multiple sheets
    with pd.ExcelWriter(excel_path) as writer:  # Use 'with' to ensure proper file closing
        
        # Create "All Songs" sheet with all tracks, sorted by album and track
        df.sort_values(['Album', 'Track']).to_excel(writer, sheet_name="All Songs", index=False)  # Sort by 'Album' and then 'Track'

        # Iterate over albums and create sheets for each one
        for album, album_df in album_groups:  # Loop through each album group
            album_df.sort_values('Track').to_excel(writer, sheet_name=album, index=False)  # Write the album DataFrame to a sheet named after the album, sorted by track

    # Load the workbook for formatting
    wb = load_workbook(excel_path)  # Load the Excel workbook

    # Apply formatting to each sheet
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))  # Define a thin border style
    for sheet_name in wb.sheetnames:  # Loop through each sheet in the workbook
        sheet = wb[sheet_name]  # Get the sheet object

        # Apply header style
        header_font = Font(bold=True)  # Define a bold font for headers
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")  # Define a gray fill for headers
        for cell in sheet[1]:  # Get the first row (header row)
            cell.font = header_font  # Apply bold font to header cells
            cell.fill = header_fill  # Apply gray fill to header cells
            cell.alignment = Alignment(horizontal='center')  # Center align the header cells

        # Auto-adjust column widths
        for column_cells in sheet.columns:  # Loop through each column
            length = max(len(str(cell.value)) for cell in column_cells)  # Get the maximum length of cell values in the column
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 2  # Set the column width to the max length + 2

        # Add gridlines
        for row in sheet.iter_rows():  # Loop through each row in the sheet
            for cell in row:  # Loop through each cell in the row
                cell.border = thin_border  # Apply the thin border to the cell

    wb.save(excel_path)  # Save the Excel file with formatting
    print(f"MP3 metadata extracted and saved to '{excel_path}' with separate sheets for each album and an 'All Songs' sheet.")

# This block of code runs only when the script is executed directly (not imported as a module)
if __name__ == "__main__":
    mp3_folder = 'G:/My Drive/MUSIC/ALBUM_REAL/'  # Specify the path to the folder containing your MP3 files
    csv_output_path = 'G:/My Drive/MUSIC/ALBUM_REAL/mp3_metadata.csv'  # Specify the path for the output CSV file
    excel_output_path = 'G:/My Drive/MUSIC/ALBUM_REAL/mp3_metadata.xlsx'  # Specify the path for the output Excel file
    extract_mp3_metadata(mp3_folder, csv_output_path, excel_output_path)  # Call the function to extract metadata and save it to files